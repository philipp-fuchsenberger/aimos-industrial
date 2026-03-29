#!/usr/bin/env python3
"""
AIMOS v4.1.0 — Shard Entrypoint
==================================
Orchestriert den AIMOSAgent-Kernel und seine Connectors.

Usage:
  python main.py                        # default agent, orchestrator mode
  python main.py --id researcher        # start agent 'researcher'
  python main.py --mode manual          # manual Telegram polling (no orchestrator)
  python main.py --debug                # verbose logging
"""

import argparse
import asyncio
import json
import logging
import os
import re
import signal
import sys
from pathlib import Path

import psutil

from core.config import Config, SecretLogFilter, make_rotating_handler
from core.agent_base import AIMOSAgent


# ── Logging Setup ─────────────────────────────────────────────────────────────

def setup_logging(agent_id: str, debug: bool = False) -> logging.Logger:
    """Configure dual logging: terminal + logs/{agent_id}.log (rotating, 10MB x 5)."""
    level = logging.DEBUG if debug else logging.INFO

    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"{agent_id}.log"

    # File handler — rotating (10MB, 5 backups)
    fh = make_rotating_handler(log_file)
    fh.setLevel(level)

    # Console handler
    fmt = logging.Formatter(
        "[%(asctime)s] %(name)-24s %(levelname)-7s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    ch = logging.StreamHandler(sys.stderr)
    ch.setLevel(level)
    ch.setFormatter(fmt)
    ch.addFilter(SecretLogFilter())

    root = logging.getLogger()
    root.setLevel(level)
    root.addHandler(fh)
    root.addHandler(ch)

    return logging.getLogger("AIMOS.main")


# ── Banner ────────────────────────────────────────────────────────────────────

def print_banner(agent_id: str, mode: str):
    log = logging.getLogger("AIMOS.main")
    log.info("=" * 60)
    log.info("  AIMOS v4.1.0 — Shard Kernel")
    log.info("=" * 60)
    log.info(f"  Agent:    {agent_id}")
    log.info(f"  Mode:     {mode}")
    log.info(f"  DB:       {Config.PG_USER}@{Config.PG_HOST}:{Config.PG_PORT}/{Config.PG_DB}")
    log.info(f"  LLM:      {Config.LLM_BASE_URL} ({Config.LLM_MODEL})")
    log.info("=" * 60)


# ── Argument Parser ───────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="AIMOS v4.1.0 — Shard Entrypoint")
    p.add_argument(
        "--id", default="agent1",
        help="Agent identifier (default: agent1)",
    )
    p.add_argument(
        "--mode", choices=["manual", "orchestrator"], default="orchestrator",
        help="manual = Telegram polling in-process; orchestrator = DB queue only (default)",
    )
    p.add_argument(
        "--debug", action="store_true",
        help="Enable DEBUG-level logging",
    )
    return p.parse_args()


# ── Main ──────────────────────────────────────────────────────────────────────

def _acquire_pidfile(agent_id: str) -> Path:
    """Singleton: create PID file or abort if agent already running."""
    pidfile = Path(f"/tmp/aimos_agent_{agent_id}.pid")
    if pidfile.exists():
        try:
            old_pid = int(pidfile.read_text().strip())
            if psutil.pid_exists(old_pid):
                # Check it's actually an AIMOS process, not a recycled PID
                try:
                    proc = psutil.Process(old_pid)
                    cmdline = " ".join(proc.cmdline())
                    if "main.py" in cmdline and agent_id in cmdline:
                        print(f"ABORT: Agent '{agent_id}' already running (PID={old_pid})")
                        sys.exit(1)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except (ValueError, OSError):
            pass
        # Stale PID file — remove it
        pidfile.unlink(missing_ok=True)
    pidfile.write_text(str(os.getpid()))
    return pidfile


def _release_pidfile(agent_id: str):
    """Remove PID file on exit."""
    Path(f"/tmp/aimos_agent_{agent_id}.pid").unlink(missing_ok=True)


async def main():
    args = parse_args()
    agent_id = args.id.lower()

    # Singleton guard
    import atexit
    pidfile = _acquire_pidfile(agent_id)
    atexit.register(_release_pidfile, agent_id)

    log = setup_logging(agent_id, debug=args.debug)
    print_banner(agent_id, args.mode)

    # ── Build agent config ────────────────────────────────────────────────
    agent_config = {
        "mode": args.mode,
        "temperature": Config.TEMPERATURE,
        "num_ctx": Config.DEFAULT_NUM_CTX,
        "history_limit": Config.HISTORY_LIMIT,
        "poll_interval": Config.POLL_INTERVAL,
    }

    agent = AIMOSAgent(agent_name=agent_id, config=agent_config)
    shutdown_event = asyncio.Event()

    # ── Signal handlers (SIGTERM + SIGINT) ────────────────────────────────
    loop = asyncio.get_running_loop()

    def _signal_handler():
        log.info("Shutdown signal received.")
        shutdown_event.set()

    for sig in (signal.SIGTERM, signal.SIGINT):
        loop.add_signal_handler(sig, _signal_handler)

    try:
        # 1. Bootstrap: DB + Schema + Seed + Secrets + Audit + History + Queue
        await agent.start()
        log.info(f"[{agent_id}] Agent bootstrapped.")

        # 2. Load tools (Brave Search etc.)
        _load_tools(agent)

        # Both modes use DB relay for Telegram (shared_listener handles sending)
        # No direct Telegram polling — prevents 409 Conflict with shared_listener
        await _run_orchestrator(agent, shutdown_event)

    except KeyboardInterrupt:
        log.info("KeyboardInterrupt — shutting down.")
    except Exception as exc:
        log.error(f"Fatal error: {exc}", exc_info=True)
    finally:
        try:
            await agent.stop()
            log.info(f"[{agent_id}] Agent stopped.")
        except Exception as exc:
            log.warning(f"Agent stop error: {exc}")


# ── Tool Loader ───────────────────────────────────────────────────────────────

def _load_tools(agent: AIMOSAgent):
    """Register tools based on agent's enabled skills from DB config.

    Error shielding: if any skill fails to load, the agent starts without it.
    """
    log = logging.getLogger("AIMOS.main")

    # Support both "skills" (v4.1+) and legacy "modules" key in DB config
    enabled = set(agent.config.get("skills", agent.config.get("modules", [])))
    if not enabled:
        enabled = {"brave_search"}

    from core.skills import SKILL_REGISTRY

    for skill_name, skill_cls in SKILL_REGISTRY.items():
        if skill_name not in enabled:
            continue
        try:
            skill = skill_cls(agent_name=agent.agent_name, agent_config=agent.config)
        except TypeError:
            # Skill doesn't accept agent_config — try agent_name only, then no args
            try:
                skill = skill_cls(agent_name=agent.agent_name, config=agent.config)
            except TypeError:
                try:
                    skill = skill_cls(agent_name=agent.agent_name)
                except TypeError:
                    try:
                        skill = skill_cls()
                    except Exception as exc:
                        log.critical(f"Skill '{skill_name}' INIT FAILED: {exc} — skipping")
                        continue
        except Exception as exc:
            log.critical(f"Skill '{skill_name}' INIT FAILED: {exc} — skipping")
            continue

        try:
            if not skill.is_available():
                log.info(f"Skill '{skill_name}' not available (missing credentials/deps)")
                continue

            for tool_def in skill.get_tools():
                tn = tool_def["name"]
                desc = tool_def.get("description", "")
                params = tool_def.get("parameters", None)

                # Create async wrapper with proper closure binding
                async def _tool_wrapper(_s=skill, _tn=tn, **kwargs):
                    return await _s.execute_tool(_tn, kwargs)
                _tool_wrapper.__doc__ = desc

                agent.register_tool(tn, _tool_wrapper, desc, parameters=params)

            log.info(f"Skill loaded: {skill_name} ({len(skill.get_tools())} tools)")
        except Exception as exc:
            log.critical(f"Skill '{skill_name}' TOOL REGISTRATION FAILED: {exc} — skipping")

    # System status tool (lightweight, no external deps)
    async def tool_system_status() -> str:
        """Zeigt Ollama-Status und VRAM-Auslastung."""
        import subprocess as _sp
        result = {"ollama": "unknown", "vram_used_mb": "?"}
        try:
            import httpx
            r = httpx.get(f"{Config.LLM_BASE_URL.rstrip('/')}/api/tags", timeout=5)
            if r.status_code == 200:
                result["ollama"] = "online"
                result["models"] = [m["name"] for m in r.json().get("models", [])]
        except Exception:
            result["ollama"] = "offline"
        try:
            r = _sp.run(["nvidia-smi", "--query-gpu=memory.used,memory.total", "--format=csv,noheader,nounits"],
                        capture_output=True, text=True, timeout=5)
            if r.returncode == 0:
                p = r.stdout.strip().split(",")
                result["vram_used_mb"] = int(p[0].strip())
                result["vram_total_mb"] = int(p[1].strip())
        except Exception:
            pass
        return json.dumps(result, indent=2)

    agent.register_tool("system_status", tool_system_status, "Zeigt Ollama-Status und VRAM")

    # Current time tool — LLMs have no clock
    async def tool_current_time() -> str:
        """Gibt das aktuelle Datum und die Uhrzeit zurueck (UTC und Lokalzeit)."""
        from datetime import datetime, timezone
        import time as _time
        utc = datetime.now(timezone.utc)
        local = datetime.now()
        tz_name = _time.tzname[_time.daylight] if _time.daylight else _time.tzname[0]
        return (
            f"UTC:   {utc.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Lokal: {local.strftime('%Y-%m-%d %H:%M:%S')} ({tz_name})\n"
            f"Wochentag: {local.strftime('%A')}"
        )
    agent.register_tool("current_time", tool_current_time, "Gibt aktuelles Datum und Uhrzeit zurueck")

    # CR-118: Send text message via Telegram to a known chat
    async def tool_send_telegram_message(text: str = "", chat_id: int = 0, **kwargs) -> str:
        """Sends a text message to a Telegram chat. If chat_id is 0, sends to the last known user.
        Use this to proactively message a user (e.g. relay information from another agent)."""
        # CR-153: Block Telegram sends when processing internal agent-to-agent messages.
        # The reply should go back to the requesting agent via dispatch_response, not to a user.
        current_kind = getattr(agent, '_current_msg_kind', '')
        if current_kind == 'internal':
            log.info(f"[{agent.agent_name}] send_telegram_message BLOCKED — internal message context. "
                     f"Reply will be relayed back to the requesting agent automatically.")
            return ("This message came from another agent, not from a user. "
                    "Your reply will be sent back to the requesting agent automatically. "
                    "Do NOT send it to Telegram — the user did not ask for this.")
        text = text or kwargs.get("message", "") or kwargs.get("content", "")
        chat_id = chat_id or kwargs.get("chat", 0)
        if not text:
            return "Error: 'text' is required."
        if not agent._pool:
            return "Error: DB not connected."
        try:
            if not chat_id:
                row = await agent._pool.fetchrow(
                    "SELECT sender_id FROM pending_messages "
                    "WHERE agent_name=$1 AND kind IN ('telegram','telegram_voice','telegram_doc') "
                    "AND sender_id IS NOT NULL AND sender_id != 0 "
                    "ORDER BY id DESC LIMIT 1",
                    agent.agent_name,
                )
                if row:
                    chat_id = int(row["sender_id"])
            if not chat_id:
                return "Error: No Telegram chat_id known. The user must message me first."
            await agent._pool.execute(
                "INSERT INTO pending_messages (agent_name, sender_id, content, kind, processed) "
                "VALUES ($1, $2, $3, 'outbound_telegram', FALSE)",
                agent.agent_name, chat_id, text,
            )
            log.info(f"[{agent.agent_name}] send_telegram_message → chat_id={chat_id}")
            return f"Message sent to Telegram chat {chat_id}."
        except Exception as exc:
            return f"Error: {exc}"

    agent.register_tool("send_telegram_message", tool_send_telegram_message,
        "Sends a text message to a Telegram user. Auto-detects chat_id from last conversation.",
        parameters={
            "text": {"type": "string", "description": "Message text to send", "required": True},
            "chat_id": {"type": "integer", "description": "Telegram chat ID (optional, auto-detected)"},
        })

    # Send file via Telegram — writes outbound_telegram_doc to DB for shared_listener
    async def tool_send_telegram_file(chat_id: int = 0, filename: str = "", caption: str = "", **kwargs) -> str:
        """Sendet eine Datei aus dem Workspace als Telegram-Dokument an einen Chat. Parameter: chat_id (Zahl), filename (Dateiname im Workspace)."""
        from pathlib import Path as _P
        # Tolerant parameter parsing — LLMs use varying names
        filename = filename or kwargs.get("file_path", "") or kwargs.get("file", "") or kwargs.get("name", "")
        chat_id = chat_id or kwargs.get("chat", 0)
        # Strip workspace prefix if LLM included full path
        if "/" in filename:
            filename = _P(filename).name
        if not filename:
            return "Error: 'filename' is required."
        if "/" in filename or "\\" in filename or ".." in filename:
            return f"Ungueltiger Dateiname: {filename}"
        workspace = _P("storage") / "agents" / agent.agent_name
        target = workspace / filename
        if not target.is_file():
            return f"Datei nicht gefunden im Workspace: {filename}"
        if not agent._pool:
            return "Error: DB not connected."
        try:
            await agent._pool.execute(
                "INSERT INTO pending_messages (agent_name, sender_id, content, kind, file_path, processed) "
                "VALUES ($1, $2, $3, 'outbound_telegram_doc', $4, FALSE)",
                agent.agent_name, int(chat_id), caption or filename, str(target),
            )
            log.info(f"[{agent.agent_name}] send_telegram_file: {filename} → chat_id={chat_id}")
            return f"File '{filename}' sent to Telegram chat {chat_id}."
        except Exception as exc:
            return f"Error: {exc}"
    agent.register_tool("send_telegram_file", tool_send_telegram_file,
        "Sendet eine Datei aus dem Workspace als Dokument per Telegram.")

    # CR-118: Send voice message via Telegram — Piper TTS → OGG → Telegram
    # Voice models: language_gender → model file. Agent picks based on context.
    _VOICE_MODELS = {
        "de_female": str(Path("models") / "leila.onnx"),
        "de_male": str(Path("models") / "thorsten.onnx"),
        "en_female": str(Path("models") / "en_amy.onnx"),
        "en_male": str(Path("models") / "en_ryan.onnx"),
        "tr_male": str(Path("models") / "tr_dfki.onnx"),
        # Aliases for convenience
        "female": str(Path("models") / "leila.onnx"),
        "male": str(Path("models") / "thorsten.onnx"),
    }
    async def tool_send_voice_message(text: str = "", voice: str = "", chat_id: int = 0, **kwargs) -> str:
        """Generates a voice message (TTS) and sends it via Telegram.
        Parameters: text (what to say), voice (e.g. 'de_female', 'de_male', 'en_female', 'en_male', 'tr_male', or just 'female'/'male' for German), chat_id (optional, auto-detected)."""
        text = text or kwargs.get("message", "") or kwargs.get("content", "")
        voice = voice or kwargs.get("gender", "")
        chat_id = chat_id or kwargs.get("chat", 0)

        if not text:
            return "Error: 'text' is required."
        if len(text) > 1000:
            text = text[:1000]  # Piper limit

        # Auto-detect voice only if not explicitly specified
        if not voice:
            # Detect language from the actual text content, not conversation history
            _lang_markers_text = {
                "tr": ["abi", "merhaba", "nasıl", "teşekkür", "lütfen", "bir", "için", "olan",
                       "değil", "benim", "senin", "yapay", "zeka", "olarak", "çok", "iyi"],
                "en": ["the", "you", "please", "thank", "would", "could", "about", "this",
                       "what", "how", "with", "from", "have", "that", "your", "are"],
                "de": ["ich", "nicht", "bitte", "danke", "guten", "kannst", "möchte",
                       "aber", "schon", "jetzt", "heute", "einen", "diese", "wird"],
            }
            text_lower = text.lower()
            scores = {lang: sum(1 for m in markers if m in text_lower)
                      for lang, markers in _lang_markers_text.items()}
            text_lang = max(scores, key=scores.get) if max(scores.values()) >= 2 else "de"

            char_desc = (agent.config.get("character", {}).get("description", "") or "").lower()
            is_male = any(w in char_desc for w in ["männlich", "male", "erkek", "kıro", "delikanlı", "abi", "er ist"])
            gender = "male" if is_male else "female"
            voice = f"{text_lang}_{gender}"
            if voice not in _VOICE_MODELS:
                voice = f"{text_lang}_male" if f"{text_lang}_male" in _VOICE_MODELS else gender
            log.info(f"[{agent.agent_name}] Voice auto-detect: text_lang={text_lang} → voice={voice}")

        voice_model = _VOICE_MODELS.get(voice.lower(), _VOICE_MODELS["female"])

        from pathlib import Path as _P
        import subprocess as _sp
        workspace = _P("storage") / "agents" / agent.agent_name
        workspace.mkdir(parents=True, exist_ok=True)
        ogg_path = workspace / f"voice_out_{int(__import__('time').time())}.ogg"

        piper_exe = "/home/philipp/AIMOS/venv/bin/piper"
        if not _P(piper_exe).exists():
            return "Error: Piper TTS not installed."
        if not _P(voice_model).exists():
            return f"Error: Voice profile not found: {voice}"

        try:
            # Piper → WAV → ffmpeg → OGG (Telegram voice format)
            wav_path = ogg_path.with_suffix(".wav")
            proc = _sp.run(
                [piper_exe, "--model", voice_model, "--output_file", str(wav_path)],
                input=text.encode("utf-8"), capture_output=True, timeout=30,
            )
            if proc.returncode != 0:
                return f"Error: Piper TTS failed: {proc.stderr.decode()[:200]}"

            # Convert WAV → OGG (Telegram requires opus/ogg for voice messages)
            _sp.run(
                ["ffmpeg", "-y", "-i", str(wav_path), "-c:a", "libopus", "-b:a", "48k", str(ogg_path)],
                capture_output=True, timeout=15,
            )
            wav_path.unlink(missing_ok=True)

            if not ogg_path.exists():
                return "Error: Audio conversion failed."

            # Find chat_id if not provided
            if not chat_id and agent._pool:
                row = await agent._pool.fetchrow(
                    "SELECT sender_id FROM pending_messages "
                    "WHERE agent_name=$1 AND kind IN ('telegram','telegram_voice','telegram_doc') "
                    "AND sender_id IS NOT NULL AND sender_id != 0 "
                    "ORDER BY id DESC LIMIT 1",
                    agent.agent_name,
                )
                if row:
                    chat_id = int(row["sender_id"])

            if not chat_id:
                return "Error: No Telegram chat known. Please provide chat_id."

            # Send via DB relay (same as send_telegram_file)
            await agent._pool.execute(
                "INSERT INTO pending_messages (agent_name, sender_id, content, kind, file_path, processed) "
                "VALUES ($1, $2, $3, 'outbound_telegram_doc', $4, FALSE)",
                agent.agent_name, chat_id, f"[Sprachnachricht] {text[:50]}...", str(ogg_path),
            )
            log.info(f"[{agent.agent_name}] send_voice_message: voice={voice} → chat_id={chat_id}")
            return f"Voice message ({voice}) sent to chat {chat_id}."
        except Exception as exc:
            return f"Error: {exc}"

    agent.register_tool("send_voice_message", tool_send_voice_message,
        "Generates a voice message (TTS) and sends it via Telegram. "
        "Available voices: de_female, de_male, en_female, en_male, tr_male. "
        "If voice is empty, auto-detects from conversation language and agent character.",
        parameters={
            "text": {"type": "string", "description": "Text to speak", "required": True},
            "voice": {"type": "string", "description": "Voice: de_female, de_male, en_female, en_male, tr_male (or empty for auto)"},
            "chat_id": {"type": "integer", "description": "Telegram chat ID (optional, auto-detected)"},
        })

    # Write file tool — create/overwrite text files in agent workspace
    async def tool_write_file(filename: str, content: str) -> str:
        """Erstellt oder ueberschreibt eine Textdatei im Workspace des Agenten."""
        from pathlib import Path as _P
        # Input validation
        if not filename or not content:
            return "Error: 'filename' and 'content' are required."
        if "/" in filename or "\\" in filename or ".." in filename:
            return f"Ungueltiger Dateiname: {filename}"
        workspace = _P("storage") / "agents" / agent.agent_name
        workspace.mkdir(parents=True, exist_ok=True)
        target = workspace / filename
        try:
            target.write_text(content, encoding="utf-8")
            size = target.stat().st_size
            log.info(f"[{agent.agent_name}] write_file: {target} ({size} bytes)")
            return f"Datei geschrieben: {filename} ({size} bytes) in {workspace}"
        except OSError as exc:
            return f"Fehler beim Schreiben: {exc}"
    agent.register_tool("write_file", tool_write_file, "Erstellt eine Textdatei im Workspace des Agenten")

    # Read file tool — read text files from agent workspace
    # CR-091: Big-File-Strategy — files >32k tokens redirect to chunking tools
    async def tool_read_file(filename: str) -> str:
        """Liest eine Textdatei aus dem Workspace des Agenten.
Bei grossen Dateien (>32k Token) wird automatisch auf den Chunking-Modus verwiesen."""
        from pathlib import Path as _P
        if not filename:
            return "Error: 'filename' is required."
        if "/" in filename or "\\" in filename or ".." in filename:
            return f"Ungueltiger Dateiname: {filename}"
        workspace = _P("storage") / "agents" / agent.agent_name
        target = workspace / filename
        if not target.is_file():
            return f"Datei nicht gefunden: {filename}"
        try:
            suffix = target.suffix.lower()

            # CR-124: Office format support — extract text from .docx, .pdf, .xlsx
            if suffix == ".docx":
                try:
                    import docx
                    doc = docx.Document(str(target))
                    text = '\n'.join(p.text for p in doc.paragraphs)
                except ImportError:
                    return "Error: python-docx is not installed."
            elif suffix == ".pdf":
                try:
                    import pdfplumber
                    with pdfplumber.open(str(target)) as pdf:
                        text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
                except ImportError:
                    return "Error: pdfplumber is not installed."
            elif suffix == ".xlsx":
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
                    lines = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        lines.append(f"--- Sheet: {sheet_name} ---")
                        for row in ws.iter_rows(values_only=True):
                            lines.append('\t'.join(str(c) if c is not None else '' for c in row))
                    wb.close()
                    text = '\n'.join(lines)
                except ImportError:
                    return "Error: openpyxl is not installed."
            else:
                text = target.read_text(encoding="utf-8", errors="replace")
            est_tokens = len(text) // 4
            # Cap at 3000 tokens (~12K chars) — leaves room for system prompt + history + response
            if est_tokens > 3000:
                chunk_size = 8_000 * 4  # chars
                total_chunks = (len(text) + chunk_size - 1) // chunk_size
                preview = text[:3000]
                return (
                    f"LARGE FILE: {filename} ({len(text)} chars, ~{est_tokens} tokens — too large for local context)\n\n"
                    f"Preview (first 3000 chars):\n{preview}\n\n"
                    f"OPTIONS to process this file:\n"
                    f"  1. read_file_chunked(filename='{filename}', chunk=0) — read chunk by chunk, summarize each\n"
                    f"  2. ask_external(question='Analyze this document', context='[paste preview above]') — let external LLM analyze\n"
                    f"  3. search_in_file(filename='{filename}', query='keyword') — search for specific content"
                )
            return text
        except OSError as exc:
            return f"Fehler beim Lesen: {exc}"
    agent.register_tool("read_file", tool_read_file, "Liest eine Textdatei aus dem Workspace")

    # ── Long-Term Memory Tools (CR-081 — Tiered Memory Architecture) ────────
    _mem_db = agent._memory_db_path

    async def tool_remember(key: str, value: str, category: str = "semantic", importance: int = 5) -> str:
        """Speichert einen Fakt dauerhaft im Langzeitgedaechtnis.

IMMER nutzen wenn:
- Der User sagt 'merk dir', 'vergiss nicht', 'wichtig'
- Persoenliche Infos geteilt werden (Namen, Vorlieben, Geburtstage)
- Entscheidungen getroffen werden die spaeter relevant sind
- Du etwas ueber den User lernst

Kategorien: semantic (Fakten/Wissen), episodic (Ereignisse), procedural (Regeln/Vorlieben)
Importance: 1-10 (10=kritisch, 7=wichtig, 5=normal, 3=nebensaechlich, 1=trivial)"""
        if not key or not value:
            return "Error: 'key' and 'value' are required."
        if not _mem_db:
            return "Error: Memory DB not initialized."
        # CR-215f: Memory write validation — prevent prompt injection via memory
        # Block system-instruction-like content and suspicious patterns
        _blocked_patterns = [
            "ignore previous", "ignore your", "system prompt", "new instructions",
            "you are now", "forget your rules", "override", "jailbreak",
            "DAN mode", "developer mode", "admin mode",
        ]
        combined = f"{key} {value}".lower()
        for pat in _blocked_patterns:
            if pat in combined:
                log.warning(f"[CR-215f] Blocked suspicious memory write: key={key[:50]}")
                return "Error: Suspicious content blocked."
        # Limit value length to prevent context stuffing
        if len(value) > 2000:
            value = value[:2000]
            log.info(f"[CR-215f] Memory value truncated to 2000 chars: key={key}")
        # Validate
        category = category.strip().lower() if category else "semantic"
        if category not in ("semantic", "episodic", "procedural"):
            category = "semantic"
        importance = max(1, min(10, int(importance) if isinstance(importance, (int, float, str)) and str(importance).isdigit() else 5))
        import sqlite3
        from core.embeddings import embed as _embed_text
        emb = _embed_text(f"{key.strip()} {value.strip()}")
        try:
            conn = sqlite3.connect(str(_mem_db), timeout=5)
            conn.execute(
                "INSERT INTO memories (key, value, category, importance, source, last_accessed, updated_at, embedding) "
                "VALUES (?, ?, ?, ?, 'user', datetime('now'), datetime('now'), ?) "
                "ON CONFLICT(key) DO UPDATE SET value=?, category=?, importance=?, "
                "updated_at=datetime('now'), last_accessed=datetime('now'), embedding=?",
                (key.strip(), value.strip(), category, importance, emb,
                 value.strip(), category, importance, emb),
            )
            # CR-140: Sync FTS5 index
            conn.execute("INSERT INTO memories_fts(memories_fts) VALUES('rebuild')")
            conn.commit()
            conn.close()
            log.info(f"[{agent.agent_name}] remember: [{category}/imp={importance}] {key} = {value[:60]}")
            return f"Stored [{category}, importance {importance}/10]: {key} = {value}"
        except Exception as exc:
            return f"Error: {exc}"
    agent.register_tool("remember", tool_remember,
        "Speichert einen Fakt dauerhaft im Langzeitgedaechtnis. Parameter: key, value, category (semantic/episodic/procedural), importance (1-10).")

    async def tool_recall(query: str = "") -> str:
        """Durchsucht das Langzeitgedaechtnis. Ohne query: zeigt die wichtigsten Eintraege. Mit query: Hybrid-Suche (Keyword + Semantik)."""
        if not _mem_db:
            return "Error: Memory DB not initialized."
        import sqlite3
        from core.embeddings import embed as _embed_text, cosine_similarity, is_available as _emb_avail
        try:
            conn = sqlite3.connect(str(_mem_db), timeout=5)
            if query and query.strip():
                q = query.strip()

                # --- Path 1: FTS5 keyword search ---
                try:
                    fts_rows = conn.execute(
                        "SELECT m.key, m.value, m.category, m.importance, m.access_count, "
                        "m.last_accessed, m.embedding "
                        "FROM memories_fts f JOIN memories m ON f.rowid = m.id "
                        "WHERE memories_fts MATCH ? ORDER BY rank LIMIT 20",
                        (q,)
                    ).fetchall()
                except Exception:
                    # FTS match syntax error — fall back to LIKE
                    like_q = f"%{q}%"
                    fts_rows = conn.execute(
                        "SELECT key, value, category, importance, access_count, "
                        "last_accessed, embedding "
                        "FROM memories WHERE key LIKE ? OR value LIKE ? "
                        "ORDER BY importance DESC LIMIT 20",
                        (like_q, like_q),
                    ).fetchall()

                # --- Path 2: Vector cosine search ---
                query_emb = _embed_text(q) if _emb_avail() else None
                vec_rows = []
                if query_emb:
                    all_rows = conn.execute(
                        "SELECT key, value, category, importance, access_count, "
                        "last_accessed, embedding "
                        "FROM memories WHERE embedding IS NOT NULL"
                    ).fetchall()
                    vec_scored = []
                    for row in all_rows:
                        if row[6]:
                            sim = cosine_similarity(query_emb, row[6])
                            vec_scored.append((sim, row))
                    vec_scored.sort(key=lambda x: x[0], reverse=True)
                    vec_rows = [r for _, r in vec_scored[:20]]

                # --- RRF Fusion (k=60) ---
                K = 60
                scores: dict[str, tuple[float, tuple]] = {}
                for rank, row in enumerate(fts_rows, start=1):
                    k = row[0]
                    prev = scores.get(k, (0.0, row[:6]))
                    scores[k] = (prev[0] + 1.0 / (K + rank), row[:6])
                for rank, row in enumerate(vec_rows, start=1):
                    k = row[0]
                    prev = scores.get(k, (0.0, row[:6]))
                    scores[k] = (prev[0] + 1.0 / (K + rank), prev[1] if prev[0] > 0 else row[:6])

                ranked = sorted(scores.values(), key=lambda x: x[0], reverse=True)
                rows = [r[1] for r in ranked[:20]]

                # Update access_count for returned results
                for r in rows:
                    conn.execute(
                        "UPDATE memories SET access_count = access_count + 1, "
                        "last_accessed = datetime('now') WHERE key = ?", (r[0],))
            else:
                rows = conn.execute(
                    "SELECT key, value, category, importance, access_count, last_accessed "
                    "FROM memories ORDER BY importance DESC, updated_at DESC LIMIT 20"
                ).fetchall()
            conn.commit()
            conn.close()
            if not rows:
                return "No entries found in long-term memory."
            lines = []
            for key, value, cat, imp, acc, la in rows[:6]:
                lines.append(f"  [{cat}|imp={imp}|x{acc}] {key}: {value}")
            return f"Long-term memory ({len(rows)} results):\n" + "\n".join(lines)
        except Exception as exc:
            return f"Error: {exc}"
    agent.register_tool("recall", tool_recall,
        "Durchsucht das Langzeitgedaechtnis nach Stichwort. Ohne query: zeigt wichtigste Eintraege.")

    async def tool_forget(key: str) -> str:
        """Loescht einen Eintrag aus dem Langzeitgedaechtnis."""
        if not key:
            return "Error: 'key' is required."
        if not _mem_db:
            return "Error: Memory DB not initialized."
        import sqlite3
        try:
            conn = sqlite3.connect(str(_mem_db), timeout=5)
            cur = conn.execute("DELETE FROM memories WHERE key = ?", (key.strip(),))
            deleted = cur.rowcount
            if deleted:
                # CR-140: Sync FTS5 index after delete
                try:
                    conn.execute("INSERT INTO memories_fts(memories_fts) VALUES('rebuild')")
                except Exception:
                    pass
            conn.commit()
            conn.close()
            if deleted:
                log.info(f"[{agent.agent_name}] forget: {key}")
                return f"Deleted: {key}"
            return f"Not found: {key}"
        except Exception as exc:
            return f"Error: {exc}"
    agent.register_tool("forget", tool_forget,
        "Loescht einen Eintrag aus dem Langzeitgedaechtnis. Parameter: key (exakter Schluessel).")

    # ── Agent-Editable Credentials (CR-083) ─────────────────────────────────
    # Agents can update credentials that are marked as "agent_editable" in config.
    # This allows e.g. an agent to store email credentials the user provides via chat.

    _editable_keys = set(agent.config.get("agent_editable_secrets", []))

    async def tool_update_credential(key: str, value: str) -> str:
        """Speichert oder aktualisiert einen Zugangsdaten-Eintrag (nur fuer freigegebene Felder).
Der Admin legt im Wizard fest, welche Felder der Agent aendern darf."""
        if not key or not value:
            return "Error: 'key' and 'value' are required."
        key = key.strip().upper()
        if key not in _editable_keys:
            allowed = ", ".join(sorted(_editable_keys)) if _editable_keys else "(none)"
            return f"Not authorized for '{key}'. Allowed fields: {allowed}"
        if not agent._pool:
            return "Error: DB not connected."
        try:
            import json as _json
            await agent._pool.execute(
                "UPDATE agents SET env_secrets = env_secrets || $1::jsonb WHERE name=$2",
                _json.dumps({key: value.strip()}), agent.agent_name,
            )
            # Also set in current process env so it takes effect immediately
            os.environ[key] = value.strip()
            log.info(f"[{agent.agent_name}] update_credential: {key} updated")
            return f"Credentials updated: {key}. Change is active immediately."
        except Exception as exc:
            return f"Error: {exc}"

    async def tool_check_credentials(**kwargs) -> str:
        """Prüft welche Zugangsdaten gesetzt sind und welche fehlen. Zeigt KEINE Werte an."""
        editable = sorted(_editable_keys) if _editable_keys else []
        # Check all known credential keys
        all_keys = sorted(set(editable) | {
            "EMAIL_ADDRESS", "EMAIL_PASSWORD", "EMAIL_IMAP_HOST", "EMAIL_SMTP_HOST",
            "TELEGRAM_BOT_TOKEN", "BRAVE_API_KEY", "OPENAI_API_KEY",
        })
        lines = []
        for key in all_keys:
            val = os.environ.get(key, "")
            status = "gesetzt" if val else "FEHLT"
            editable_flag = " (änderbar)" if key in _editable_keys else " (nur Admin)"
            lines.append(f"  {key}: {status}{editable_flag}")
        return "Credential-Status:\n" + "\n".join(lines)

    agent.register_tool("check_credentials", tool_check_credentials,
        "Zeigt welche Zugangsdaten gesetzt sind und welche fehlen (ohne Werte)")

    if _editable_keys:
        agent.register_tool("update_credential", tool_update_credential,
            f"Aktualisiert Zugangsdaten. Erlaubte Felder: {', '.join(sorted(_editable_keys))}")
        log.info(f"Agent-editable credentials: {sorted(_editable_keys)}")

    # ── CR-065/CR-096: Internal Messenger — agent-to-agent communication ─────
    # Configurable: config.inter_agent_messaging = true/false (default: true)
    # Optional whitelist: config.allowed_agents = ["agent1", "agent2"] (empty = all allowed)
    _iam_enabled = agent.config.get("inter_agent_messaging", True)
    _allowed_agents = [a.strip().lower() for a in agent.config.get("allowed_agents", []) if a.strip()]

    async def tool_send_to_agent(agent_name: str = "", message: str = "", context_file: str = "", **kwargs) -> str:
        """Sendet eine Nachricht an einen anderen AIMOS-Agenten.
        Der Ziel-Agent erhält die Nachricht in seiner Queue und wird geweckt.
        Optional: context_file — Dateiname aus deinem Workspace der als Kontext mitgesendet wird.
        Parameter: agent_name (Zielname), message (Nachrichtentext), context_file (optional)."""
        agent_name = agent_name or kwargs.get("target", "") or kwargs.get("name", "")
        message = message or kwargs.get("content", "") or kwargs.get("text", "")
        context_file = context_file or kwargs.get("file", "")

        if not agent_name or not message:
            return "Error: 'agent_name' and 'message' are required."

        target = agent_name.strip().lower()

        if target == agent.agent_name:
            return "Error: Cannot send messages to yourself."

        # CR-096: Whitelist check
        if _allowed_agents and target not in _allowed_agents:
            return f"Error: Not authorized to send to '{target}'. Allowed: {', '.join(_allowed_agents)}"

        if not agent._pool or agent._pool._closed:
            return "Error: DB not connected."

        try:
            async with agent._pool.acquire() as conn:
                exists = await conn.fetchval(
                    "SELECT 1 FROM agents WHERE LOWER(name)=$1", target
                )
                if not exists:
                    return f"Error: Agent '{target}' does not exist."

                # CR-148: Attach workspace file as context if specified
                full_message = f"[Nachricht von {agent.agent_name}] {message}"
                if context_file:
                    context_file = context_file.strip().replace("..", "").lstrip("/")
                    from core.skills.base import BaseSkill
                    ctx_path = BaseSkill.workspace_path(agent.agent_name) / context_file
                    if ctx_path.exists() and ctx_path.is_file():
                        try:
                            ctx_content = ctx_path.read_text(encoding="utf-8", errors="replace")[:2000]
                            full_message += f"\n\n[Attached: {context_file}]\n{ctx_content}"
                        except Exception:
                            pass

                # CR-thread: Propagate thread_id to internal messages
                _thread_id = getattr(agent, '_current_thread_id', '') or ''

                # Auto-replace customer data in delegation message from Kundenakte.
                # The LLM hallucinates placeholder names/addresses — we replace
                # the ENTIRE message with verified data from the customer file.
                if _thread_id:
                    import json as _json_sta
                    _cust_dir_sta = Path("storage/customers")
                    if _cust_dir_sta.exists():
                        for _cf_sta in _cust_dir_sta.glob("*.json"):
                            try:
                                _cd_sta = _json_sta.loads(_cf_sta.read_text(encoding="utf-8"))
                                if _thread_id in _cd_sta.get("thread_ids", []):
                                    _cust_name = _cd_sta.get("name", "")
                                    _cust_company = _cd_sta.get("company", "")
                                    _cust_email = _cd_sta.get("email", "")
                                    _cust_addr = _cd_sta.get("address", "")
                                    _cust_products = ", ".join(_cd_sta.get("products", []))
                                    # Replace entire message — keep only the task instruction,
                                    # strip all LLM-hallucinated customer names/addresses
                                    import re as _re_sta
                                    # Extract just the action verb/intent from the LLM message
                                    _task = _re_sta.sub(
                                        r'(?:Kunde|Customer|Herr|Frau|Name|Firma|Adresse|Email|Lieferadresse|Bestellung an)[:\s]+[^.!\n]*[.!\n]?',
                                        '', message, flags=_re_sta.IGNORECASE
                                    ).strip()
                                    if not _task or len(_task) < 10:
                                        _task = message  # fallback: keep original if stripping removed everything
                                    full_message = (
                                        f"[Nachricht von {agent.agent_name}] "
                                        f"Kunde: {_cust_name}, Firma: {_cust_company}, "
                                        f"Email: {_cust_email}, Adresse: {_cust_addr}. "
                                        f"Produkte: {_cust_products}. "
                                        f"Aufgabe: {_task}"
                                    )
                                    break
                            except Exception:
                                pass

                msg_id = await conn.fetchval(
                    "INSERT INTO pending_messages (agent_name, sender_id, content, kind, thread_id) "
                    "VALUES ($1, 0, $2, 'internal', $3) RETURNING id",
                    target, full_message, _thread_id,
                )
                await conn.execute(
                    "UPDATE agents SET wake_up_needed=TRUE WHERE LOWER(name)=$1", target
                )

            agent._delegated_this_cycle = True
            log.info(f"[{agent.agent_name}] send_to_agent: #{msg_id} → '{target}'")
            return f"Message #{msg_id} sent to '{target}'."
        except Exception as exc:
            return f"Error: {exc}"

    # CR-061: Read files from another agent's public/ folder
    async def tool_read_public(agent_name: str = "", filename: str = "", **kwargs) -> str:
        """Read a file from another agent's public folder. Agents can share files
        by placing them in their public/ directory. Parameter: agent_name, filename."""
        agent_name = agent_name or kwargs.get("name", "")
        filename = filename or kwargs.get("file", "")
        if not agent_name or not filename:
            return "Error: 'agent_name' and 'filename' are required."
        target = agent_name.strip().lower()
        filename = filename.strip().replace("..", "").lstrip("/")
        from core.skills.base import BaseSkill
        public_path = BaseSkill.workspace_path(target) / "public" / filename
        if not public_path.exists():
            # List available public files
            public_dir = BaseSkill.workspace_path(target) / "public"
            if public_dir.exists():
                files = [f.name for f in public_dir.iterdir() if f.is_file()]
                return f"File '{filename}' not found in {target}/public/. Available: {', '.join(files) or '(empty)'}"
            return f"Agent '{target}' has no public folder."
        try:
            content = public_path.read_text(encoding="utf-8", errors="replace")
            if len(content) > 3000:
                content = content[:3000] + "\n[... truncated, use read_file_chunked for full content]"
            return content
        except Exception as exc:
            return f"Error reading: {exc}"

    agent.register_tool("read_public", tool_read_public,
        "Read a file from another agent's public/ folder. For cross-agent file sharing.",
        parameters={"agent_name": {"type": "string", "description": "Agent name"},
                     "filename": {"type": "string", "description": "Filename in public/ folder"}})

    # ── Thread Lookup — cross-channel thread linking ────────────────────────
    async def tool_lookup_thread(customer_name: str = "", email: str = "",
                                  subject: str = "", **kwargs) -> str:
        """Sucht einen bestehenden Thread (z.B. Email-Konversation) anhand von
        Kundenname, E-Mail-Adresse oder Betreff. Gibt die letzten Nachrichten
        aus dem gefundenen Thread zurueck. Nutze dieses Tool wenn ein Kunde
        auf Telegram behauptet, eine Email geschrieben zu haben.
        Parameter: customer_name, email, subject (mindestens eins angeben)."""
        if not agent._pool:
            return "DB nicht verfuegbar."
        if not any([customer_name, email, subject]):
            return "At least customer_name, email, or subject required."
        try:
            # Search threads by content match
            conditions = []
            params = [agent.agent_name]
            idx = 2
            if email:
                conditions.append(f"LOWER(content) LIKE LOWER(${idx})")
                params.append(f"%{email.strip()}%")
                idx += 1
            if customer_name:
                conditions.append(f"LOWER(content) LIKE LOWER(${idx})")
                params.append(f"%{customer_name.strip()}%")
                idx += 1
            if subject:
                conditions.append(f"LOWER(content) LIKE LOWER(${idx})")
                params.append(f"%{subject.strip()}%")
                idx += 1
            where = " OR ".join(conditions)
            rows = await agent._pool.fetch(
                f"SELECT DISTINCT thread_id, content, created_at FROM aimos_chat_histories "
                f"WHERE agent_name=$1 AND thread_id LIKE 'email:%' AND ({where}) "
                f"ORDER BY created_at DESC LIMIT 5",
                *params,
            )
            if not rows:
                return (f"Kein Email-Thread gefunden fuer: "
                        f"name={customer_name or '-'}, email={email or '-'}, "
                        f"betreff={subject or '-'}. "
                        f"Moegliche Ursachen: Thread existiert nicht, oder "
                        f"die Angaben stimmen nicht ueberein.")
            # Return thread info + verification data
            found_threads = {}
            for r in rows:
                tid = r["thread_id"]
                if tid not in found_threads:
                    found_threads[tid] = []
                found_threads[tid].append(r["content"][:300])

            result_parts = []
            for tid, msgs in found_threads.items():
                # Load full thread context
                thread_msgs = await agent._pool.fetch(
                    "SELECT role, content, created_at FROM aimos_chat_histories "
                    "WHERE agent_name=$1 AND thread_id=$2 "
                    "ORDER BY created_at DESC LIMIT 10",
                    agent.agent_name, tid,
                )
                result_parts.append(f"=== Thread {tid} ({len(thread_msgs)} Nachrichten) ===")
                for m in reversed(thread_msgs):
                    result_parts.append(f"[{m['role']}] {(m['content'] or '')[:400]}")

            # Verification hint
            result_parts.append(
                "\n--- VERIFIKATION ---\n"
                "Pruefe ob Name/Email aus dem Thread mit den Angaben des "
                "Telegram-Kunden uebereinstimmen. Bei Abweichungen: "
                "hoeflich nachfragen."
            )
            return "\n".join(result_parts)
        except Exception as exc:
            return f"Thread-Suche fehlgeschlagen: {exc}"

    agent.register_tool("lookup_thread", tool_lookup_thread,
        "Sucht einen bestehenden Email-Thread anhand von Kundenname, "
        "E-Mail-Adresse oder Betreff. Fuer Cross-Channel Verknuepfung.",
        parameters={
            "customer_name": {"type": "string", "description": "Kundenname"},
            "email": {"type": "string", "description": "E-Mail-Adresse"},
            "subject": {"type": "string", "description": "Betreff oder Stichwort"},
        })

    # ── CR-231: Structured customer files (Kundenakte) ──────────────────────
    async def tool_update_customer(
        name: str = "", company: str = "", email: str = "",
        phone: str = "", address: str = "",
        product: str = "", order: str = "", note: str = "",
        **kwargs
    ) -> str:
        """Updates the customer file (Kundenakte). Call this whenever you learn
        new information about a customer. Parameters (all optional, only pass what's new):
        name, company, email, phone, address, product (add to product list),
        order (e.g. '2x JUNIOR II — Angebot gesendet'), note (free text)."""
        name = name or kwargs.get("name", "")
        if not name:
            return "Error: 'name' is required."
        if len(name) > 200:
            return "Error: customer name too long (max 200 chars)."

        import json as _json_cust
        import re as _re_cust
        from datetime import datetime as _dt_cust
        import fcntl

        # Slugify customer_id from name — strict whitelist, no path traversal possible
        _slug = _re_cust.sub(r'[^a-z0-9]+', '-', name.lower().strip()).strip('-')[:100]
        if not _slug or '/' in _slug or '..' in _slug:
            return "Error: invalid customer name."

        cust_dir = Path("storage/customers")
        cust_dir.mkdir(parents=True, exist_ok=True)

        # Dedup: search existing files for this customer before creating a new one
        # "Brandner" should find "karl-brandner.json" and update it
        cust_file = cust_dir / f"{_slug}.json"
        if not cust_file.exists():
            # Search all existing files for matching name/email
            _name_lower = name.lower()
            _variants = {_name_lower}
            for u, a in [("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss")]:
                if u in _name_lower: _variants.add(_name_lower.replace(u, a))
            for a, u in [("ae", "ä"), ("oe", "ö"), ("ue", "ü")]:
                if a in _name_lower: _variants.add(_name_lower.replace(a, u))
            for existing in cust_dir.glob("*.json"):
                try:
                    _ex_data = _json_cust.loads(existing.read_text(encoding="utf-8"))
                    _ex_searchable = _json_cust.dumps(_ex_data, ensure_ascii=False).lower()
                    # Check if any variant of the name appears in existing file
                    if any(v in _ex_searchable for v in _variants):
                        cust_file = existing
                        _slug = existing.stem
                        log.info(f"[CR-231] Matched existing customer file: {existing.name} for '{name}'")
                        break
                    # Also match by slug substring
                    if any(v in existing.stem for v in _variants) or existing.stem in _slug:
                        cust_file = existing
                        _slug = existing.stem
                        log.info(f"[CR-231] Matched existing customer file by slug: {existing.name}")
                        break
                except Exception:
                    pass

        _lock_file = cust_dir / f".{_slug}.lock"

        # File-locked read-modify-write to prevent race conditions
        try:
            with open(_lock_file, "w") as _lf:
                fcntl.flock(_lf, fcntl.LOCK_EX)

                # Load existing or create new (with JSON error recovery)
                if cust_file.exists():
                    try:
                        data = _json_cust.loads(cust_file.read_text(encoding="utf-8"))
                    except _json_cust.JSONDecodeError:
                        log.warning(f"[CR-231] Corrupted customer file {_slug}.json — recreating")
                        data = None
                else:
                    data = None

                if data is None:
                    data = {
                        "name": name,
                        "company": "",
                        "email": "",
                        "phone": "",
                        "address": "",
                        "products": [],
                        "orders": [],
                        "notes": [],
                        "thread_ids": [],
                        "created": _dt_cust.now().strftime("%Y-%m-%d %H:%M"),
                        "last_contact": "",
                    }

                # Update fields (only if provided and non-empty)
                if name:
                    data["name"] = name
                if company:
                    data["company"] = company
                if email:
                    data["email"] = email
                if phone:
                    data["phone"] = phone
                if address:
                    data["address"] = address
                if product and product not in data.get("products", []):
                    data.setdefault("products", []).append(product)
                if order:
                    _order_entry = f"{_dt_cust.now().strftime('%d.%m.')} {order}"
                    data.setdefault("orders", []).append(_order_entry)
                if note:
                    _note_entry = f"{_dt_cust.now().strftime('%d.%m.')} {note}"
                    data.setdefault("notes", []).append(_note_entry)
                    data["notes"] = data["notes"][-10:]

                # Link current thread_id if available
                _cur_tid = getattr(agent, '_current_thread_id', '')
                if _cur_tid and _cur_tid not in data.get("thread_ids", []):
                    data.setdefault("thread_ids", []).append(_cur_tid)

                data["last_contact"] = _dt_cust.now().strftime("%Y-%m-%d %H:%M")

                # Atomic write: write to temp, then rename
                _tmp = cust_file.with_suffix(".tmp")
                _tmp.write_text(_json_cust.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                _tmp.rename(cust_file)

                fcntl.flock(_lf, fcntl.LOCK_UN)
        except Exception as _cust_exc:
            return f"Error updating customer file: {_cust_exc}"

        return f"Customer file updated: {_slug}.json"

    agent.register_tool("update_customer", tool_update_customer,
        "Updates the customer file (Kundenakte). Call whenever you learn new info about a customer.",
        parameters={
            "name": {"type": "string", "description": "Customer name (required)"},
            "company": {"type": "string", "description": "Company name"},
            "email": {"type": "string", "description": "Email address"},
            "phone": {"type": "string", "description": "Phone number"},
            "address": {"type": "string", "description": "Address"},
            "product": {"type": "string", "description": "Product to add to list"},
            "order": {"type": "string", "description": "Order entry (e.g. '2x JUNIOR II — Angebot gesendet')"},
            "note": {"type": "string", "description": "Free text note"},
        })

    _tools_list = "system_status, current_time, write_file, read_file, remember, recall, forget, read_public, lookup_thread, update_customer"
    if _iam_enabled:
        agent.register_tool("send_to_agent", tool_send_to_agent,
            "Sendet eine Nachricht an einen anderen AIMOS-Agenten"
            + (f" (erlaubt: {', '.join(_allowed_agents)})" if _allowed_agents else ""))
        _tools_list += ", send_to_agent"

    log.info(f"System tools registered: {_tools_list}"
             + (f", update_credential" if _editable_keys else ""))


# ── Voice Transcription ──────────────────────────────────────────────────────

_VOICE_FAIL_MSG = (
    "Ich habe die Sprachnachricht empfangen, konnte sie aber "
    "technisch nicht entschlüsseln. Bitte sende den Text schriftlich."
)


def _detect_language_from_history(agent, log) -> str:
    """CR-116: Detect expected language from recent chat history.

    Scans the last 5 assistant messages for language indicators.
    Returns an ISO 639-1 language code (de, en, tr, fr, es, ar, etc.)
    or 'auto' if uncertain.
    """
    # Language detection heuristics — keywords that strongly indicate a language
    _LANG_MARKERS = {
        "tr": ["abi", "moruk", "merhaba", "teşekkür", "evet", "hayır", "tamam", "nasıl",
               "günaydın", "selam", "şeker", "lütfen", "galatasaray", "değil"],
        "de": ["ich", "nicht", "bitte", "danke", "guten", "morgen", "kannst", "möchte",
               "wir", "aber", "schon", "jetzt", "heute", "gestern"],
        "en": ["the", "you", "please", "thank", "would", "could", "should", "about",
               "what", "where", "when", "how", "this", "that"],
        "fr": ["je", "vous", "merci", "bonjour", "oui", "non", "comment", "pourquoi"],
        "es": ["hola", "gracias", "por favor", "cómo", "qué", "estoy", "muy"],
        "ar": ["مرحبا", "شكرا", "نعم", "لا", "كيف", "ماذا"],
    }

    try:
        # Only look at user messages on the SAME channel (not internal agent chatter)
        recent = agent._history[-20:] if len(agent._history) >= 20 else agent._history
        text_blob = " ".join(
            e.get("content", "")[:200].lower()
            for e in recent
            if e.get("role") in ("assistant", "user")
            and "channel=internal" not in e.get("content", "")
            and "[Nachricht von" not in e.get("content", "")
        )

        if not text_blob:
            return agent.config.get("whisper_language", "auto")

        # Count marker hits per language
        scores = {}
        for lang, markers in _LANG_MARKERS.items():
            scores[lang] = sum(1 for m in markers if m in text_blob)

        best_lang = max(scores, key=scores.get) if scores else "auto"
        best_score = scores.get(best_lang, 0)

        if best_score >= 3:
            log.info(f"[{agent.agent_name}] Voice language detected: {best_lang} (score={best_score})")
            return best_lang

        # Not confident — use config default or auto
        return agent.config.get("whisper_language", "auto")

    except Exception:
        return agent.config.get("whisper_language", "auto")


async def _transcribe_voice(agent, voice_path: str, log) -> str:
    """Transcribe a voice file via Whisper with auto-detected language.

    CR-116: Detects expected language from conversation history before
    transcribing. Supports any language Whisper knows — not limited to
    pre-configured languages.

    Returns transcribed text, or a fixed error message on failure.
    """
    from pathlib import Path

    path = Path(voice_path)
    if not path.is_file():
        log.error(f"[{agent.agent_name}] Voice file not found: {voice_path}")
        return _VOICE_FAIL_MSG

    # CR-116: Detect language from conversation context
    detected_lang = _detect_language_from_history(agent, log)
    log.info(f"[{agent.agent_name}] Voice: transcribing {path.name} (language={detected_lang})")

    # CR-131: Whisper runs on CPU now — no VRAM flush needed.
    # (CR-117 VRAM flush removed — was causing 10s model reload overhead every voice message)

    try:
        from core.skills.voice_io import VoiceIOSkill
        vio = VoiceIOSkill()

        # CR-116 fix: ALWAYS let Whisper auto-detect language first (medium model is good at this).
        # Forcing a language causes Whisper to "translate" instead of transcribe.
        # Only use language hint as secondary validation, not as primary choice.
        candidates = [None]  # None = Whisper auto-detect (ALWAYS first)

        best_text = ""
        best_lang = "auto"

        for lang in candidates:
            vio._whisper_language = lang
            text = await vio.transcribe_file(str(path))

            if not text or not text.strip():
                continue

            # Quality check: reject gibberish (too many repeated chars, too short, all symbols)
            clean = text.strip()
            words = clean.split()
            unique_chars = len(set(clean.lower()))

            # Good transcription: enough words, enough unique characters, not all punctuation
            is_good = (
                len(words) >= 2
                and unique_chars >= 5
                and sum(1 for c in clean if c.isalpha()) > len(clean) * 0.5
            )

            if is_good:
                best_text = clean
                best_lang = lang
                log.info(f"[{agent.agent_name}] Voice: accepted transcription (lang={lang}): '{clean[:80]}'")
                break
            else:
                log.info(f"[{agent.agent_name}] Voice: rejected transcription (lang={lang}): '{clean[:60]}' — trying next")
                if not best_text:
                    best_text = clean  # keep as last resort
                    best_lang = lang

        if not best_text:
            log.warning(f"[{agent.agent_name}] Voice: all transcription attempts failed")
            return _VOICE_FAIL_MSG

        log.info(f"[{agent.agent_name}] Voice: final result (lang={best_lang}): '{best_text[:80]}'")

        try:
            path.unlink()
        except OSError:
            pass

        return f"[Sprachnachricht transkribiert] {best_text}"

    except ImportError as exc:
        log.error(f"[{agent.agent_name}] Voice: faster-whisper not installed: {exc}")
        return _VOICE_FAIL_MSG
    except Exception as exc:
        log.error(f"[{agent.agent_name}] Voice: transcription failed: {exc}")
        return _VOICE_FAIL_MSG


# ── Auto-Remember Safety Net ─────────────────────────────────────────────────

_REMEMBER_TRIGGERS = re.compile(
    r"(?:"
    # German
    r"merk\s*dir|vergiss\s*nicht|wichtig.*merken|notier\s*dir|"
    r"speicher\s*(?:dir|das)|behalte?\s*(?:das|im\s*kopf)|"
    # English
    r"remember\s+(?:this|that)|don'?t\s+forget|keep\s+in\s+mind|note\s+(?:this|that)|"
    # Turkish
    r"bunu\s*(?:not\s+et|unutma|kaydet|hat[ıi]rla)|sakla|aklında\s+tut|"
    # French
    r"souviens[\s-]*toi|n'?oublie\s*pas|note[\s-]*(?:le|ça)|retiens|"
    # Spanish
    r"recuerda\s+(?:esto|que)|no\s+olvides|anota|ten\s+en\s+cuenta"
    r")",
    re.IGNORECASE,
)


async def _auto_remember(agent, user_content: str, reply: str, log):
    """Safety net: if user explicitly asked to remember something but agent didn't
    call the remember tool, extract the fact and store it automatically.

    Checks the audit log for a recent TOOL_START remember entry.
    """
    if not _REMEMBER_TRIGGERS.search(user_content):
        return  # user didn't ask to remember anything

    # Check if remember was actually called (via audit log)
    if agent._audit_path and agent._audit_path.exists():
        try:
            lines = agent._audit_path.read_text().strip().split("\n")
            recent = lines[-5:] if len(lines) >= 5 else lines
            if any("TOOL_START" in l and "remember" in l for l in recent):
                return  # tool was called — no intervention needed
        except OSError:
            pass

    # Agent didn't call remember — extract and store automatically
    # Use the user's message as the value, generate a key from content
    clean = re.sub(r"\[Kontext:[^\]]*\]\n?", "", user_content).strip()
    if len(clean) < 5:
        return

    # Simple key extraction: first meaningful words
    words = re.sub(r"[^a-zA-ZäöüßÄÖÜ0-9\s]", "", clean).split()
    key_words = [w for w in words if len(w) > 2 and w.lower() not in
                 ("merk", "dir", "dass", "merke", "bitte", "vergiss", "nicht",
                  "wichtig", "merken", "auch", "noch", "und", "der", "die",
                  "das", "ein", "eine", "wir", "haben", "ist", "sind")][:4]
    key = "_".join(w.lower() for w in key_words) if key_words else "auto_memory"

    if not agent._memory_db_path:
        return
    import sqlite3
    from core.embeddings import embed as _embed_text
    emb = _embed_text(f"{key} {clean}")
    try:
        conn = sqlite3.connect(str(agent._memory_db_path), timeout=5)
        conn.execute(
            "INSERT INTO memories (key, value, category, importance, source, last_accessed, updated_at, embedding) "
            "VALUES (?, ?, 'semantic', 7, 'auto', datetime('now'), datetime('now'), ?) "
            "ON CONFLICT(key) DO UPDATE SET value=?, importance=7, updated_at=datetime('now'), "
            "last_accessed=datetime('now'), embedding=?",
            (key, clean, emb, clean, emb),
        )
        # CR-140: Sync FTS5 index
        try:
            conn.execute("INSERT INTO memories_fts(memories_fts) VALUES('rebuild')")
        except Exception:
            pass
        conn.commit()
        conn.close()
        log.info(f"[{agent.agent_name}] AUTO-REMEMBER: {key} = {clean[:60]}")
    except Exception as exc:
        log.warning(f"[{agent.agent_name}] auto-remember failed: {exc}")


# ── Auto-Followup Safety Net ────────────────────────────────────────────────

_FOLLOWUP_RE = re.compile(
    r"(?:soll ich|m[oö]chten?\s+(?:Sie|du)|warten\s+(?:Sie|wir)|"
    r"lass(?:en)?\s+(?:Sie\s+)?(?:mich|es|uns)\s+wissen|"
    r"gib(?:st|t)?\s+(?:mir|uns)\s+(?:bitte\s+)?bescheid|"
    r"ich\s+warte\s+auf|ich\s+hake\s+nach|"
    r"bei\s+(?:R[uü]ck|Interesse|Bedarf|Fragen)|"
    r"melde(?:st|t|n)?\s+(?:dich|Sie\s+sich)|"
    r"ich\s+(?:bleibe|bin)\s+dran)",
    re.IGNORECASE,
)


async def _external_fallback(agent, user_message: str, msg: dict, log,
                              reason: str = "timeout") -> str:
    """Escalate to external API when local LLM fails. Returns LLM-generated reply.

    Three phases:
    1. Dispatch an interim "working on it" message (LLM-generated in user's language)
    2. Get the actual answer from external API with natural lead-in
    3. Dream-on-demand: extract facts from history, store in memory, clear history
       → frees context for the local LLM on the next message

    All text is LLM-generated — no hardcoded language fragments.
    """
    try:
        from core.skills.skill_hybrid_reasoning import HybridReasoningSkill
        hr = HybridReasoningSkill(agent.agent_name, agent.config)
        if not hr.is_available():
            log.warning(f"[{agent.agent_name}] External API not available for fallback")
            return ""

        # Build conversation summary for context
        history_summary = ""
        try:
            recent = agent._history[-10:] if hasattr(agent, '_history') else []
            if recent:
                history_lines = []
                for m in recent:
                    role = m.get("role", "?")
                    content = (m.get("content", ""))[:200]
                    if content and not content.startswith("Tool "):
                        history_lines.append(f"{role}: {content}")
                history_summary = "\n".join(history_lines[-6:])
        except Exception:
            pass

        # Phase 1+2 combined: try to get the answer directly.
        # Only send "please wait" if the answer takes >5 seconds.
        answer_context = (
            f"You are answering on behalf of a specialized support agent. "
            f"The local AI could not process this question (reason: {reason}). "
            f"Provide a helpful answer. "
            f"Respond in the SAME LANGUAGE as the user's question. "
            f"Be concise (max 200 words). Do not use emojis. Be professional."
            f"\n\nRecent conversation for context:\n{history_summary}" if history_summary else
            f"You are answering on behalf of a specialized support agent. "
            f"The local AI could not process this question (reason: {reason}). "
            f"Provide a helpful answer. "
            f"Respond in the SAME LANGUAGE as the user's question. "
            f"Be concise (max 200 words). Do not use emojis. Be professional."
        )
        answer_task = asyncio.create_task(hr.execute_tool("ask_external", {
            "question": user_message[:800],
            "context": answer_context,
        }))

        # Wait up to 5s for the answer — if it's not ready, send interim
        INTERIM_DELAY = 5
        interim_sent = False
        try:
            answer = await asyncio.wait_for(asyncio.shield(answer_task), timeout=INTERIM_DELAY)
        except asyncio.TimeoutError:
            # Answer taking long — send "please wait" now
            try:
                interim = await hr.execute_tool("ask_external", {
                    "question": user_message[:200],
                    "context": (
                        "The user asked this question but the system needs a moment to research. "
                        "Generate ONLY a brief 1-sentence polite 'please wait' message in the SAME "
                        "LANGUAGE as the user's question. Do not answer the question itself yet. "
                        "Do not use emojis. Example style: 'One moment please, I am researching this for you.'"
                    ),
                })
                if interim and len(interim.strip()) > 5:
                    # Skip interim for email — "please wait" makes no sense asynchronously
                    _msg_kind = msg.get("kind", "")
                    if _msg_kind != "email":
                        await agent.dispatch_response(interim.strip(), msg)
                        interim_sent = True
                        log.info(f"[{agent.agent_name}] Fallback Phase 1: interim dispatched (answer took >{INTERIM_DELAY}s)")
                    else:
                        log.info(f"[{agent.agent_name}] Fallback Phase 1: skipped for email (async channel)")
            except Exception as exc:
                log.debug(f"[{agent.agent_name}] Interim message failed: {exc}")
            # Now wait for the actual answer
            answer = await answer_task

        if answer and len(answer.strip()) > 20:
            log.info(f"[{agent.agent_name}] Fallback Phase 2: answer received ({len(answer)} chars)")

            # Phase 3: Dream-on-demand — extract facts + clear history
            await _emergency_dream(agent, hr, log)

            return answer.strip()

    except Exception as exc:
        log.error(f"[{agent.agent_name}] External fallback failed completely: {exc}")

    return ""


async def _emergency_dream(agent, hr, log):
    """Extract key facts from conversation via external API, store in memory, clear history.

    This frees context for the local LLM. Runs after an escalation to external API.
    """
    try:
        # Collect recent conversation
        recent = agent._history[-15:] if hasattr(agent, '_history') else []
        if len(recent) < 4:
            return  # Not enough history to dream about

        conv_lines = []
        for m in recent:
            role = m.get("role", "?")
            content = (m.get("content", ""))[:300]
            if content and not content.startswith("Tool "):
                conv_lines.append(f"{role}: {content}")

        if not conv_lines:
            return

        conversation_text = "\n".join(conv_lines)

        # Ask external API to extract facts
        extraction = await hr.execute_tool("ask_external", {
            "question": "Extract key facts from this conversation.",
            "context": (
                "Extract the most important facts from this conversation as a list. "
                "Each fact should be one line in the format: KEY: VALUE\n"
                "Examples:\n"
                "customer_name: Thomas Mueller, Feuerwehr Rosenheim\n"
                "equipment: 3x CAPITANO II\n"
                "serial_BK-2024-22222: 890 operating hours\n"
                "issue_reported: error E-03, low oil pressure\n\n"
                "Only extract facts worth remembering for future conversations. "
                "Max 10 facts. No commentary, just KEY: VALUE lines.\n\n"
                f"CONVERSATION:\n{conversation_text}"
            ),
        })

        if not extraction or len(extraction.strip()) < 10:
            return

        # Parse and store facts in agent's SQLite memory
        import sqlite3
        from core.skills.base import BaseSkill
        db_path = BaseSkill.memory_db_path(agent.agent_name)
        if not db_path.exists():
            return

        conn = sqlite3.connect(str(db_path), timeout=5)
        stored = 0
        for line in extraction.strip().split("\n"):
            line = line.strip()
            if ":" not in line or len(line) < 5:
                continue
            key, _, value = line.partition(":")
            key = key.strip().lower().replace(" ", "_")[:100]
            value = value.strip()[:500]
            if not key or not value:
                continue
            try:
                conn.execute(
                    "INSERT OR REPLACE INTO memories (key, value, category, importance, updated_at) "
                    "VALUES (?, ?, 'semantic', 7, datetime('now'))",
                    (f"dream_{key}", value)
                )
                stored += 1
            except Exception:
                pass
        conn.commit()
        conn.close()

        # Clear chat history to free context
        if stored > 0 and agent._pool:
            deleted = await agent._pool.fetchval(
                "WITH excess AS ("
                "  SELECT id FROM aimos_chat_histories "
                "  WHERE agent_name=$1 "
                "  AND id NOT IN ("
                "    SELECT id FROM aimos_chat_histories WHERE agent_name=$1 "
                "    ORDER BY id DESC LIMIT 4"
                "  )"
                ") "
                "DELETE FROM aimos_chat_histories WHERE id IN (SELECT id FROM excess) "
                "RETURNING id",
                agent.agent_name,
            )
            # Also clear in-memory history
            if hasattr(agent, '_history') and len(agent._history) > 4:
                agent._history = agent._history[-4:]

            log.info(
                f"[{agent.agent_name}] Emergency dream: {stored} facts stored, "
                f"{deleted or 0} history entries cleared"
            )

    except Exception as exc:
        log.warning(f"[{agent.agent_name}] Emergency dream failed (non-critical): {exc}")


def _merge_queued_messages(messages: list[dict], agent, log) -> list[list[dict]]:
    """CR-206: Merge queued messages from the same sender+channel into batches.

    Natural chat behavior: user sends "Hello" + "I need help with X" quickly.
    Instead of 2 separate LLM calls, merge into one: "Hello\nI need help with X".

    Grouping rules:
    - Same sender_id AND same kind → merge
    - Different sender or kind → separate batch
    - Max ~4000 chars per batch (leave room in context)
    - Each batch becomes one think() call

    Returns list of message groups. Each group is a list of msg dicts.
    """
    if not messages:
        return []

    # CR-211: Sort by sender wait time (oldest first = fairest)
    # Messages from different senders get interleaved fairly
    messages.sort(key=lambda m: m.get("created_at") or "")

    MAX_BATCH_CHARS = 4000  # Leave room for system prompt + memory + history
    ESCALATE_THRESHOLD = 8000  # If total queued content > this, escalate to external API

    batches = []
    current_batch = []
    current_chars = 0
    current_key = None  # (sender_id, kind)

    for msg in messages:
        sender_id = msg.get("sender_id", 0)
        kind = msg.get("kind", "text")
        content = msg.get("content", "")

        # Group Telegram family together (text, voice, doc, photo all from same user)
        kind_family = kind
        if kind in ("telegram", "telegram_voice", "telegram_doc", "telegram_photo"):
            kind_family = "telegram_family"
        # L3: Include thread_id in merge key — emails all have sender_id=0,
        # so without thread_id they'd wrongly merge into one batch
        key = (sender_id, kind_family, msg.get("thread_id", ""))

        if key != current_key or current_chars + len(content) > MAX_BATCH_CHARS:
            # Start new batch
            if current_batch:
                batches.append(current_batch)
            current_batch = [msg]
            current_chars = len(content)
            current_key = key
        else:
            # Add to current batch
            current_batch.append(msg)
            current_chars += len(content)

    if current_batch:
        batches.append(current_batch)

    # Check total volume — if too much for local LLM, mark for escalation
    total_chars = sum(len(m.get("content", "")) for m in messages)
    if total_chars > ESCALATE_THRESHOLD:
        log.warning(
            f"[CR-206] Queue overflow: {total_chars} chars in {len(messages)} messages "
            f"→ exceeds local context budget. Marking batches for external escalation."
        )
        # Mark each batch for escalation
        for batch in batches:
            batch[0]["_escalate_to_external"] = True

    # Log merges
    for batch in batches:
        if len(batch) > 1:
            log.info(
                f"[CR-206] Merged {len(batch)} messages: "
                f"sender={batch[0].get('sender_id')}, kind={batch[0].get('kind')}, "
                f"total_chars={sum(len(m.get('content', '')) for m in batch)}"
            )

    return batches


async def _auto_followup(agent, reply: str, log):
    """Safety net: if the agent's reply implies it's waiting for a response
    or has open action items, but didn't set a reminder — create one automatically.

    Only fires if no set_reminder was called in this cycle (check audit log).
    """
    if not _FOLLOWUP_RE.search(reply):
        return  # reply doesn't imply waiting

    # Check if set_reminder was already called
    if agent._audit_path and agent._audit_path.exists():
        try:
            lines = agent._audit_path.read_text().strip().split("\n")
            recent = lines[-8:] if len(lines) >= 8 else lines
            if any("TOOL_START" in l and "set_reminder" in l for l in recent):
                return  # agent already set a reminder
        except OSError:
            pass

    # Auto-set a 30-minute followup (CR-110: max 3 auto-followups per agent)
    topic = reply[:80].replace("\n", " ").strip()
    try:
        import asyncpg
        from core.config import Config
        conn = await asyncpg.connect(**Config.get_db_params())

        # CR-110: Don't pile up auto-followups
        pending_followups = await conn.fetchval(
            "SELECT COUNT(*) FROM agent_jobs WHERE agent_name=$1 AND source='auto_followup' AND status='pending'",
            agent.agent_name,
        )
        if pending_followups >= 3:
            await conn.close()
            log.info(f"[{agent.agent_name}] AUTO-FOLLOWUP: skipped (already {pending_followups} pending)")
            return

        from datetime import datetime, timezone, timedelta
        scheduled = datetime.now(timezone.utc) + timedelta(minutes=30)
        await conn.execute(
            "INSERT INTO agent_jobs (agent_name, scheduled_time, task_prompt, source) "
            "VALUES ($1, $2, $3, 'auto_followup')",
            agent.agent_name, scheduled,
            f"[Auto-Followup] Du hattest auf eine Antwort gewartet zum Thema: {topic}. "
            f"Pruefe ob der User geantwortet hat und hake ggf. nach.",
        )
        await conn.close()
        log.info(f"[{agent.agent_name}] AUTO-FOLLOWUP: scheduled in 30m — {topic[:50]}")
    except Exception as exc:
        log.warning(f"[{agent.agent_name}] auto-followup failed: {exc}")


# ── Mode: Orchestrator (DB queue polling, replies via DB relay) ────────────────

async def _run_orchestrator(agent: AIMOSAgent, shutdown_event: asyncio.Event):
    """Orchestrator mode: process pending_messages, route replies via dispatch_response."""
    log = logging.getLogger("AIMOS.main")
    log.info(f"[{agent.agent_name}] Orchestrator mode — DB queue, dispatch_response routing.")

    async def _process_loop():
        poll_interval = agent.config.get("poll_interval", Config.POLL_INTERVAL)
        if agent._pool:
            await agent._pool.execute(
                "UPDATE agents SET status='running', updated_at=NOW() WHERE name=$1",
                agent.agent_name,
            )

        watchdog_task = asyncio.create_task(agent._watchdog())
        try:
            while not shutdown_event.is_set():
                messages = await agent.poll_pending()
                if messages:
                    agent._touch()  # reset watchdog on ANY activity
                    if agent._pool:
                        await agent._pool.execute(
                            "UPDATE agents SET status='active', updated_at=NOW() WHERE name=$1",
                            agent.agent_name,
                        )
                else:
                    # No messages — set status to 'idle' (watchdog will fire after 600s)
                    if agent._pool:
                        await agent._pool.execute(
                            "UPDATE agents SET status='idle', updated_at=NOW() WHERE name=$1",
                            agent.agent_name,
                        )
                # CR-206: Merge queued messages from same sender+channel into one request
                # This matches natural chat behavior (user sends multiple messages quickly)
                merged_messages = _merge_queued_messages(messages, agent, log)

                for msg_group in merged_messages:
                    msg = msg_group[0]  # Primary message (for routing metadata)
                    content = msg.get("content", "")
                    sender_id = msg.get("sender_id", 0)
                    kind = msg.get("kind", "text")

                    # CR-203: Skip scheduled jobs if agent has disable_auto_jobs
                    if kind == "scheduled_job" and agent.config.get("disable_auto_jobs"):
                        log.info(f"[{agent.agent_name}] Skipping scheduled_job (disable_auto_jobs=true)")
                        continue
                    voice_path = msg.get("file_path", "")

                    # If multiple messages were merged, combine their content
                    if len(msg_group) > 1:
                        combined_parts = []
                        for m in msg_group:
                            mc = m.get("content", "")
                            mk = m.get("kind", "text")
                            mvp = m.get("file_path", "")
                            # Voice transcription
                            if mk == "telegram_voice" and mvp:
                                mc = await _transcribe_voice(agent, mvp, log)
                            elif mk == "telegram_doc" and mvp:
                                from pathlib import Path as _P
                                fname = _P(mvp).name
                                mc = f"[Dokument empfangen: {fname}] Nutze read_file(filename=\"{fname}\") um sie zu lesen."
                            if mc.strip():
                                combined_parts.append(mc.strip())
                        content = "\n".join(combined_parts)
                        log.info(
                            f"[{agent.agent_name}] MERGED {len(msg_group)} messages from "
                            f"sender={sender_id} kind={kind}: {len(content)} chars total"
                        )
                    else:
                        log.info(
                            f"[{agent.agent_name}] MSG id={msg.get('id')} "
                            f"sender={sender_id} kind={kind} len={len(content)}"
                        )
                        # Voice transcription: download → Whisper → text
                        if kind == "telegram_voice" and voice_path:
                            content = await _transcribe_voice(agent, voice_path, log)
                        # CR-118: Document received
                        if kind == "telegram_doc" and voice_path:
                            from pathlib import Path as _P
                            fname = _P(voice_path).name
                            content = f"[Dokument empfangen: {fname}] Nutze read_file(filename=\"{fname}\") um sie zu lesen."

                    # Inject sender context + timestamp
                    ts = msg.get("created_at")
                    if ts:
                        from datetime import timezone as _tz
                        if hasattr(ts, "astimezone"):
                            local_ts = ts.astimezone().strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            local_ts = str(ts)[:19]
                    else:
                        from datetime import datetime as _dt
                        local_ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

                    # CR-123: Conversation Partner Awareness
                    if kind == "telegram":
                        sender_name = "Helpdesk"
                    elif kind == "internal":
                        sender_name = "Agent"
                    elif kind == "scheduled_job":
                        sender_name = "System (Cronjob)"
                    else:
                        sender_name = "System"

                    # Skip empty content (e.g. emoji-only message after CJK filter)
                    if not content or not content.strip():
                        log.info(f"[{agent.agent_name}] Skipping empty message (id={msg.get('id')})")
                        continue

                    # CR-153: Track message kind so tools can check context
                    agent._current_msg_kind = kind
                    agent._last_msg_content = content  # For send_to_agent reply-to extraction

                    # CR-209: Set session_id for multi-user history isolation
                    # CR-thread: Set thread_id from message (propagated from shared_listener)
                    # IMPORTANT: Must be set BEFORE customer context injection below
                    msg_thread_id = msg.get("thread_id", "") or ""
                    if kind in ("telegram", "telegram_voice", "telegram_doc"):
                        agent._current_session_id = f"telegram:{sender_id}"
                        agent._current_thread_id = msg_thread_id or f"tg:{sender_id}"
                    elif kind == "email":
                        agent._current_session_id = f"email:{sender_id}"
                        agent._current_thread_id = msg_thread_id or f"email:{sender_id}"
                    elif kind == "internal":
                        agent._current_session_id = f"internal:{sender_id}"
                        agent._current_thread_id = msg_thread_id or f"internal:{sender_id}"
                    else:
                        agent._current_session_id = f"dashboard:{msg.get('id', 0)}"
                        agent._current_thread_id = msg_thread_id or f"dashboard:{msg.get('id', 0)}"

                    # Inject current customer context from Kundenakte if available
                    _customer_hint = ""
                    _cur_thread = getattr(agent, '_current_thread_id', '') or ''
                    if _cur_thread:
                        import json as _json_ctx
                        _cust_dir = Path("storage/customers")
                        if _cust_dir.exists():
                            for _cf in _cust_dir.glob("*.json"):
                                try:
                                    _cd = _json_ctx.loads(_cf.read_text(encoding="utf-8"))
                                    if _cur_thread in _cd.get("thread_ids", []):
                                        _customer_hint = (
                                            f"\n[Current customer: {_cd.get('name', '?')}"
                                            f" | {_cd.get('company', '')}"
                                            f" | Email: {_cd.get('email', '?')}"
                                            f" | Products: {', '.join(_cd.get('products', []))}"
                                            f" | Orders: {'; '.join(_cd.get('orders', [])[-2:])}"
                                            f"]\n[IMPORTANT: This conversation is ONLY about this customer. "
                                            f"Ignore any memory entries about other customers.]"
                                        )
                                        break
                                except Exception:
                                    pass

                    content_with_ctx = (
                        f"[Von: {sender_name} | Kanal: {kind} | Zeit: {local_ts}]"
                        f"{_customer_hint}\n{content}"
                    )

                    # CR-156: Per-message tool-call budget
                    agent._tool_call_count = 0
                    agent._tool_call_budget = agent.config.get("max_tool_calls_per_message", 10)

                    # CR-206: If marked for escalation (queue overflow), go directly to external API
                    if msg.get("_escalate_to_external"):
                        log.info(f"[{agent.agent_name}] CR-206: Queue overflow → direct external API escalation")
                        reply = await _external_fallback(agent, content, msg, log,
                                                         reason="queue_overflow")
                        if reply:
                            # Skip think() entirely — external API handled it
                            agent._touch()
                            await _auto_followup(agent, reply, log) if not agent.config.get("disable_auto_jobs") else None
                            route = await agent.dispatch_response(reply, msg)
                            log.info(f"[{agent.agent_name}] Dispatched (external) → {route}")
                            continue

                    # CR-207b removed: Image analysis is now handled by the agent's
                    # analyze_image tool directly (no code-level bypass needed).

                    try:
                        reply = await asyncio.wait_for(agent.think(content_with_ctx), timeout=120)
                        log.info(f"[{agent.agent_name}] think() → {len(reply)} chars")
                        # If response filter stripped everything → escalate to external API
                        if not reply or not reply.strip():
                            log.warning(f"[{agent.agent_name}] think() returned empty — escalating to external API")
                            reply = await _external_fallback(agent, content, msg, log,
                                                             reason="empty_response")
                            msg["_fallback_handled"] = True
                    except asyncio.TimeoutError:
                        log.error(f"[{agent.agent_name}] think() TIMED OUT — escalating to external API")
                        reply = await _external_fallback(agent, content, msg, log,
                                                         reason="timeout")
                        msg["_fallback_handled"] = True
                    except Exception as exc:
                        log.error(f"[{agent.agent_name}] think() FAILED: {exc}")
                        reply = await _external_fallback(agent, content, msg, log,
                                                         reason=f"error: {exc}")
                        msg["_fallback_handled"] = True

                    agent._touch()

                    # CR-213: Check-before-Send (config: check_before_send=true)
                    # After think(), check if more messages arrived from same sender.
                    # If so, discard reply, merge everything, re-process.
                    # Config: image_delegate — agent name to route images to first.
                    if (
                        agent.config.get("check_before_send")
                        and kind not in ("internal", "scheduled_job")
                    ):
                        late_msgs = await agent.poll_pending()
                        if not late_msgs:
                            log.info(f"[CR-213] Check-before-Send: no late messages — sending reply")
                        if late_msgs:
                            # L4: For email kind, filter by thread_id instead of sender_id
                            # (sender_id=0 for all emails would match everything)
                            if kind == "email":
                                _current_thread = getattr(agent, '_current_thread_id', '')
                                related_msgs = [m for m in late_msgs if m.get("thread_id", "") == _current_thread]
                                other_msgs = [m for m in late_msgs if m.get("thread_id", "") != _current_thread]
                            else:
                                related_msgs = [m for m in late_msgs if m.get("sender_id") == sender_id]
                                other_msgs = [m for m in late_msgs if m.get("sender_id") != sender_id]
                            # Alias for backward compat in the block below
                            same_sender = related_msgs
                            if same_sender:
                                log.info(
                                    f"[CR-213] Check-before-Send: {len(same_sender)} late message(s) "
                                    f"from sender={sender_id} — discarding reply, will re-process"
                                )
                                # Merge original + late messages, re-process as one
                                all_reprocess = list(msg_group) + same_sender
                                combined_parts = []
                                for m in all_reprocess:
                                    mc = m.get("content", "")
                                    mk = m.get("kind", "text")
                                    mvp = m.get("file_path", "")
                                    if mk == "telegram_voice" and mvp:
                                        mc = await _transcribe_voice(agent, mvp, log)
                                    elif mk == "telegram_doc" and mvp:
                                        from pathlib import Path as _P
                                        fname = _P(mvp).name
                                        mc = f"[Dokument empfangen: {fname}] Nutze read_file(filename=\"{fname}\") um sie zu lesen."
                                    if mc and mc.strip():
                                        combined_parts.append(mc.strip())
                                merged_content = "\n".join(combined_parts)
                                content_with_ctx = (
                                    f"[Von: {sender_name} | Kanal: {kind} | Zeit: {local_ts}]\n"
                                    f"{merged_content}"
                                )
                                log.info(
                                    f"[CR-213] Re-processing {len(all_reprocess)} messages "
                                    f"as one ({len(merged_content)} chars)"
                                )
                                try:
                                    reply = await asyncio.wait_for(
                                        agent.think(content_with_ctx), timeout=120
                                    )
                                    log.info(f"[{agent.agent_name}] think() retry → {len(reply)} chars")
                                    if not reply or not reply.strip():
                                        reply = await _external_fallback(
                                            agent, merged_content, msg, log,
                                            reason="empty_response_retry"
                                        )
                                except (asyncio.TimeoutError, Exception) as exc:
                                    log.error(f"[{agent.agent_name}] think() retry failed: {exc}")
                                    reply = await _external_fallback(
                                        agent, merged_content, msg, log,
                                        reason=f"retry_error: {exc}"
                                    )

                            # Re-queue messages from other senders
                            if other_msgs and agent._pool:
                                for om in other_msgs:
                                    await agent._pool.execute(
                                        "INSERT INTO pending_messages "
                                        "(agent_name, sender_id, content, kind, thread_id) "
                                        "VALUES ($1, $2, $3, $4, $5)",
                                        agent.agent_name,
                                        om.get("sender_id", 0),
                                        om.get("content", ""),
                                        om.get("kind", "text"),
                                        om.get("thread_id", ""),
                                    )
                                log.info(
                                    f"[CR-213] Re-queued {len(other_msgs)} message(s) "
                                    f"from other senders"
                                )
                    # CR-145: Update contact's last_interaction timestamp
                    try:
                        from core.skills.skill_contacts import auto_update_contact_from_message
                        auto_update_contact_from_message(agent.agent_name, msg.get("sender_id", 0), kind)
                    except Exception:
                        pass

                    # Auto-remember: REMOVED — dreaming Phase 0 (LLM-powered) handles
                    # memory extraction from conversations much more precisely than
                    # regex triggers, which caused false positives.

                    # Auto-followup: if agent's reply implies waiting, set a reminder
                    # BUT NOT for scheduled_job replies — that creates infinite loops
                    # (cronjob fires → agent says "waiting" → auto-followup → new cronjob → repeat)
                    if kind not in ("scheduled_job", "internal"):
                        if not agent.config.get("disable_auto_jobs"):
                            await _auto_followup(agent, reply, log)

                    # Route reply via agent's dispatch_response (decoupled from main.py)
                    route = await agent.dispatch_response(reply, msg)
                    log.info(f"[{agent.agent_name}] Dispatched → {route}")

                    # CR-thread: Auto-notify after email send
                    # If the agent called send_email during this think cycle,
                    # notify the configured agent (via notify_on_email config)
                    notify_target = agent.config.get("notify_on_email", "")
                    if (notify_target
                            and getattr(agent, '_email_sent_this_cycle', False)
                            and agent._pool
                            and agent.agent_name != notify_target):
                        try:
                            _notify_thread_id = getattr(agent, '_current_thread_id', '') or ''
                            await agent._pool.execute(
                                "INSERT INTO pending_messages "
                                "(agent_name, sender_id, content, kind, thread_id) "
                                "VALUES ($1, 0, $2, 'internal', $3)",
                                notify_target,
                                f"[Nachricht von {agent.agent_name}] "
                                f"E-Mail wurde gesendet. Antwort des Agenten: {reply[:500]}",
                                _notify_thread_id,
                            )
                            await agent._pool.execute(
                                "UPDATE agents SET wake_up_needed=TRUE WHERE LOWER(name)=$1",
                                notify_target,
                            )
                            log.info(
                                f"[CR-thread] Auto-notify: {agent.agent_name} sent email "
                                f"→ notified {notify_target} (thread={_notify_thread_id})"
                            )
                        except Exception as exc:
                            log.warning(f"[CR-thread] Auto-notify failed: {exc}")

                await asyncio.sleep(poll_interval)
        except asyncio.CancelledError:
            pass
        finally:
            watchdog_task.cancel()

    loop_task = asyncio.create_task(_process_loop())
    shutdown_task = asyncio.create_task(shutdown_event.wait())

    done, pending = await asyncio.wait(
        [loop_task, shutdown_task],
        return_when=asyncio.FIRST_COMPLETED,
    )
    for task in pending:
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass


# ── Entry ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass
